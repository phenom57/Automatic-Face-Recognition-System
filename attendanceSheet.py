import xlrd
import xlwt
from xlutils.copy import copy
import os.path
import openpyxl
from datetime import date

subject = ["AI", "Cryptography", "DP", "DIP", "ADBMS"]
wbName = 'attendance.xlsx'
wb = openpyxl.load_workbook(wbName)
sheet = wb['ADBMS']
# r_sheet = rb.sheet_by_index(subject.index("Distributed Processing"))
today = date.today().isoformat()
# print(r_sheet)
# print(r_sheet.nrows)
# print(r_sheet.ncols)

# sheet.cell(row = 1, column = sheet.max_column).value = today
print(sheet.max_row)
print(sheet.max_column)
# print(range(r_sheet.ncolumns))
# for i in range(r_sheet.nrows) :
# 	print (r_sheet.row_values(i));
for i in range(1, sheet.max_column) :
	print(sheet.cell(row = 1, column = i).value)
wb.save('attendance.xlsx')
wb.close
# wb = copy(rb)
# sheet = wb.get_sheet(0)
# sheet.write(0, 1, datetime.now().date())
# wb.save('attendance.xls')
# for i in worksheet : 
# 	print(i)

# worksheet.write(1, 0, datetime.now())

# workbook.close()