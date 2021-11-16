import cv2
import numpy as np
import sqlite3
import os
from gtts import gTTS
import openpyxl
from datetime import date
import pickle

conn = sqlite3.connect('database.db')
c = conn.cursor()

fname = "recognizer/trainingData.yml"

if not os.path.isfile(fname):
  print("Please train the data first")
  exit(0)

face_classifier = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
cap = cv2.VideoCapture(0)

recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read(fname)
num = 0
name = ''
subject = ["AI", "Cryptography", "DP", "DIP", "ADBMS"]
finalName = '';
wbName = 'attendance.xlsx'
studentId = 0
while True:
  ret, frame = cap.read()
  gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
  equ = cv2.equalizeHist(gray)
  final = cv2.medianBlur(equ, 3)

  faces = face_classifier.detectMultiScale(gray, 1.3, 5)
  for (x,y,w,h) in faces:
    cv2.rectangle(frame,(x,y),(x+w,y+h),(0, 255, 0), 3)
    ids,conf = recognizer.predict(gray[y:y+h,x:x+w])
    c.execute("select name from Student where rollNo = (?);", (ids,))
    result = c.fetchall()
    name = result[0][0]
    if len(name) != 0 :
      studentId = ids
      finalName = name;
    
    if conf < 50:
      cv2.putText(frame, name, (x + 2, y + h - 4), cv2.FONT_HERSHEY_SIMPLEX, 1, (150, 255, 0), 2)
    else:
      cv2.putText(frame, 'No Match', (x+2,y+h-5), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,0,255),2)
  
  cv2.imshow('Face Recognizer',frame)
  # Esc key for closing frame
  K = cv2.waitKey(30) & 0xff == 27
  num += 1
  
  if num > 50 :
    textAttendance = ''
    if len(finalName) != 0 :
      textAttendance = finalName + 'Your Attendance has been registered for today'
    else :
      textAttendance = 'Kindly retry to give the Attendance'
    tts = gTTS(textAttendance)
    tts.save("register.mp3")
    
    f = open('subject.pckl', 'rb')
    subject = pickle.load(f)
    wb = openpyxl.load_workbook(wbName)
    sheet = wb[subject]
    today = date.today().isoformat()
    col = -1
    len = sheet.max_column + 1
    for i in range(1, len):
      if sheet.cell(row = 1, column = i).value == today :
        col = i;
        break

    if col == -1 :
      col = sheet.max_column + 1

    sheet.cell(row = 1, column = col).value = today
    sheet.cell(row = studentId + 1, column = col).value = "Present"
    wb.save('attendance.xlsx')
    wb.close()
    os.system("mpg321 register.mp3")
    cap.release()
    break

cap.release()
cv2.destroyAllWindows()
conn.commit()
conn.close()